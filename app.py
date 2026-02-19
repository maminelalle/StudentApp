from datetime import date

from flask import (
    Flask,
    render_template,
    redirect,
    url_for,
    request,
    flash,
    session,
)
from sqlalchemy import (
    create_engine,
    Column,
    Integer,
    String,
    Date,
    Float,
    ForeignKey,
    func,
    select,
)
from sqlalchemy.orm import declarative_base, relationship, sessionmaker, scoped_session


Base = declarative_base()


class Filiere(Base):
    __tablename__ = "filiere"

    id = Column(Integer, primary_key=True)
    nom = Column(String, nullable=False, unique=True)  # L1, L2, L3
    description = Column(String)

    etudiants = relationship("Etudiant", back_populates="filiere")
    specialites = relationship("Specialite", back_populates="filiere")
    matieres = relationship("Matiere", back_populates="filiere")


class Specialite(Base):
    __tablename__ = "specialite"

    id = Column(Integer, primary_key=True)
    nom = Column(String, nullable=False)  # DSI, RSS, CNM
    code = Column(String, nullable=False)  # DSI, RSS, CNM
    description = Column(String)
    filiere_id = Column(Integer, ForeignKey("filiere.id"), nullable=False)

    filiere = relationship("Filiere", back_populates="specialites")
    etudiants = relationship("Etudiant", back_populates="specialite")
    matieres = relationship("Matiere", back_populates="specialite")


class Semestre(Base):
    __tablename__ = "semestre"

    id = Column(Integer, primary_key=True)
    nom = Column(String, nullable=False, unique=True)  # S1, S2, S3, S4, S5, S6
    numero = Column(Integer, nullable=False)  # 1, 2, 3, 4, 5, 6

    matieres = relationship("Matiere", back_populates="semestre")


class Matiere(Base):
    __tablename__ = "matiere"

    id = Column(Integer, primary_key=True)
    code = Column(String, nullable=False, unique=True)  # Ex: INFO101
    nom = Column(String, nullable=False)
    semestre_id = Column(Integer, ForeignKey("semestre.id"), nullable=True)
    filiere_id = Column(Integer, ForeignKey("filiere.id"), nullable=True)
    specialite_id = Column(Integer, ForeignKey("specialite.id"), nullable=True)
    seuil_validation = Column(Float, default=10.0)
    rattrapable = Column(Integer, default=1)  # 1 = oui, 0 = non

    semestre = relationship("Semestre", back_populates="matieres")
    filiere = relationship("Filiere", back_populates="matieres")
    specialite = relationship("Specialite", back_populates="matieres")
    devoirs = relationship("Devoir", back_populates="matiere", cascade="all, delete-orphan")


class Devoir(Base):
    __tablename__ = "devoir"

    id = Column(Integer, primary_key=True)
    nom = Column(String, nullable=False)
    date = Column(Date, nullable=False)
    type = Column(String, nullable=False, default="Écrit")  # TP, Écrit, Projet, Examen, Rattrapage
    session = Column(String, nullable=False, default="Normale")  # Normale, Rattrapage
    matiere_id = Column(Integer, ForeignKey("matiere.id"), nullable=False)

    matiere = relationship("Matiere", back_populates="devoirs")
    notes = relationship("Note", back_populates="devoir", cascade="all, delete-orphan")


class Etudiant(Base):
    __tablename__ = "etudiant"

    id = Column(Integer, primary_key=True)
    nom = Column(String, nullable=False)
    email = Column(String, nullable=False, unique=True)
    mot_de_passe = Column(String, nullable=False)
    filiere_id = Column(Integer, ForeignKey("filiere.id"), nullable=True)
    specialite_id = Column(Integer, ForeignKey("specialite.id"), nullable=True)

    filiere = relationship("Filiere", back_populates="etudiants")
    specialite = relationship("Specialite", back_populates="etudiants")
    notes = relationship("Note", back_populates="etudiant", cascade="all, delete-orphan")


class Utilisateur(Base):
    __tablename__ = "utilisateur"

    id = Column(Integer, primary_key=True)
    nom = Column(String, nullable=False)
    email = Column(String, nullable=False, unique=True)
    mot_de_passe = Column(String, nullable=False)
    role = Column(String, nullable=False)  # admin, enseignant


class Note(Base):
    __tablename__ = "note"

    id = Column(Integer, primary_key=True)
    etudiant_id = Column(Integer, ForeignKey("etudiant.id"), nullable=False)
    devoir_id = Column(Integer, ForeignKey("devoir.id"), nullable=False)
    valeur = Column(Float, nullable=False)

    etudiant = relationship("Etudiant", back_populates="notes")
    devoir = relationship("Devoir", back_populates="notes")


def create_app():
    app = Flask(__name__)
    app.config["SECRET_KEY"] = "change-this-secret-key"
    app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///student_grades.db"

    engine = create_engine(app.config["SQLALCHEMY_DATABASE_URI"], echo=False, future=True)
    Base.metadata.create_all(engine)

    # Seed : données initiales si la base est vide
    with engine.connect() as conn:
        from sqlalchemy import text
        
        # Vérifier si les données existent déjà
        r = conn.execute(text("SELECT COUNT(*) FROM utilisateur")).scalar()
        if r == 0:
            # Créer admin
            conn.execute(text(
                "INSERT INTO utilisateur (nom, email, mot_de_passe, role) VALUES "
                "('Admin', 'admin@studentapp.com', 'admin123', 'admin')"
            ))
            
            # Créer filières
            conn.execute(text(
                "INSERT INTO filiere (nom, description) VALUES "
                "('L1', 'Licence 1'), "
                "('L2', 'Licence 2'), "
                "('L3', 'Licence 3')"
            ))
            
            # Créer spécialités (toutes pour L3 principalement)
            conn.execute(text(
                "INSERT INTO specialite (nom, code, description, filiere_id) VALUES "
                "('Développement Système Informatique', 'DSI', 'Spécialisation en développement logiciel et systèmes', 3), "
                "('Réseau et Sécurité', 'RSS', 'Spécialisation en réseaux et cybersécurité', 3), "
                "('Communication et Numérique Multimédia', 'CNM', 'Spécialisation en création multimédia', 3)"
            ))
            
            # Créer semestres
            conn.execute(text(
                "INSERT INTO semestre (nom, numero) VALUES "
                "('S1', 1), ('S2', 2), ('S3', 3), ('S4', 4), ('S5', 5), ('S6', 6)"
            ))
            
            # Créer étudiants de test
            conn.execute(text(
                "INSERT INTO etudiant (nom, email, mot_de_passe, filiere_id, specialite_id) VALUES "
                "('Étudiant Demo', 'etudiant@studentapp.com', 'etudiant123', 3, 1), "
                "('Alice Martin', 'alice.martin@supnum.tn', 'alice123', 3, 1), "
                "('Bob Dupont', 'bob.dupont@supnum.tn', 'bob123', 3, 2), "
                "('Clara Benali', 'clara.benali@supnum.tn', 'clara123', 2, NULL), "
                "('David Touati', 'david.touati@supnum.tn', 'david123', 1, NULL)"
            ))
            
            # Créer matières de test
            conn.execute(text(
                "INSERT INTO matiere (code, nom, semestre_id, filiere_id, specialite_id, seuil_validation, rattrapable) VALUES "
                # L1 - S1
                "('MATH101', 'Mathématiques 1', 1, 1, NULL, 10.0, 1), "
                "('INFO101', 'Introduction à l''informatique', 1, 1, NULL, 10.0, 1), "
                "('PHY101', 'Physique 1', 1, 1, NULL, 10.0, 1), "
                # L1 - S2
                "('MATH102', 'Mathématiques 2', 2, 1, NULL, 10.0, 1), "
                "('INFO102', 'Algorithmes et structures de données', 2, 1, NULL, 10.0, 1), "
                # L3 DSI - S5
                "('DSI501', 'Développement Web Avancé', 5, 3, 1, 10.0, 1), "
                "('DSI502', 'Base de Données Avancées', 5, 3, 1, 10.0, 1), "
                "('DSI503', 'Génie Logiciel', 5, 3, 1, 10.0, 1), "
                "('DSI504', 'Programmation Mobile', 5, 3, 1, 10.0, 1), "
                # L3 RSS - S5
                "('RSS501', 'Sécurité des Réseaux', 5, 3, 2, 10.0, 1), "
                "('RSS502', 'Administration Systèmes', 5, 3, 2, 10.0, 1), "
                "('RSS503', 'Cryptographie', 5, 3, 2, 10.0, 1)"
            ))
            
            # Créer devoirs de test
            conn.execute(text(
                "INSERT INTO devoir (nom, date, type, session, matiere_id) VALUES "
                # Matières L1
                "('TP1 - Variables et Opérations', '2026-02-25', 'TP', 'Normale', 2), "
                "('DS1 - Logique', '2026-03-05', 'Écrit', 'Normale', 2), "
                "('Projet - Application Console', '2026-03-20', 'Projet', 'Normale', 2), "
                "('Examen Final Info', '2026-06-15', 'Examen', 'Normale', 2), "
                # Matières L3 DSI
                "('TP1 - React Basics', '2026-02-20', 'TP', 'Normale', 6), "
                "('TP2 - REST API', '2026-03-01', 'TP', 'Normale', 6), "
                "('Projet Web', '2026-04-15', 'Projet', 'Normale', 6), "
                "('Examen Final Web', '2026-06-10', 'Examen', 'Normale', 6), "
                "('TP1 - SQL Avancé', '2026-02-22', 'TP', 'Normale', 7), "
                "('DS1 - Requêtes', '2026-03-10', 'Écrit', 'Normale', 7), "
                "('Projet BDD', '2026-04-20', 'Projet', 'Normale', 7), "
                "('TP1 - UML', '2026-02-18', 'TP', 'Normale', 8), "
                "('DS1 - Conception', '2026-03-08', 'Écrit', 'Normale', 8), "
                "('Projet GL', '2026-05-01', 'Projet', 'Normale', 8), "
                "('TP1 - Android Studio', '2026-02-24', 'TP', 'Normale', 9), "
                "('Projet Mobile', '2026-04-25', 'Projet', 'Normale', 9), "
                # Matières L3 RSS
                "('TP1 - Firewall', '2026-02-21', 'TP', 'Normale', 10), "
                "('DS1 - Protocoles', '2026-03-12', 'Écrit', 'Normale', 10), "
                "('Projet Sécurité', '2026-04-18', 'Projet', 'Normale', 10), "
                "('TP1 - Linux Admin', '2026-02-23', 'TP', 'Normale', 11), "
                "('DS1 - Shell', '2026-03-15', 'Écrit', 'Normale', 11), "
                "('TP1 - RSA', '2026-02-26', 'TP', 'Normale', 12), "
                "('DS1 - Chiffrement', '2026-03-18', 'Écrit', 'Normale', 12)"
            ))
            
            # Créer notes de test pour les étudiants
            # Étudiant 1 (Demo) - L3 DSI
            conn.execute(text(
                "INSERT INTO note (etudiant_id, devoir_id, valeur) VALUES "
                "(1, 5, 15.5), (1, 6, 14.0), (1, 7, 16.5), (1, 8, 13.0), "  # Dev Web
                "(1, 9, 17.0), (1, 10, 15.5), (1, 11, 14.5), "  # BDD
                "(1, 12, 16.0), (1, 13, 15.0), (1, 14, 17.5), "  # Génie Log
                "(1, 15, 14.0), (1, 16, 15.5)"  # Mobile
            ))
            
            # Étudiant 2 (Alice) - L3 DSI
            conn.execute(text(
                "INSERT INTO note (etudiant_id, devoir_id, valeur) VALUES "
                "(2, 5, 18.0), (2, 6, 17.5), (2, 7, 19.0), (2, 8, 16.5), "
                "(2, 9, 18.5), (2, 10, 17.0), (2, 11, 18.0), "
                "(2, 12, 17.5), (2, 13, 16.5), (2, 14, 19.0), "
                "(2, 15, 16.0), (2, 16, 17.5)"
            ))
            
            # Étudiant 3 (Bob) - L3 RSS
            conn.execute(text(
                "INSERT INTO note (etudiant_id, devoir_id, valeur) VALUES "
                "(3, 17, 14.5), (3, 18, 13.0), (3, 19, 15.0), "  # Sécurité
                "(3, 20, 16.5), (3, 21, 15.0), "  # Admin
                "(3, 22, 17.0), (3, 23, 16.0)"  # Crypto
            ))
            
            # Étudiant 4 (Clara) - L2 (pas de notes encore)
            
            # Étudiant 5 (David) - L1
            conn.execute(text(
                "INSERT INTO note (etudiant_id, devoir_id, valeur) VALUES "
                "(5, 1, 12.5), (5, 2, 11.0), (5, 3, 13.5), (5, 4, 12.0)"
            ))
            
            conn.commit()

    SessionLocal = scoped_session(sessionmaker(bind=engine, autoflush=False, autocommit=False))

    def get_db():
        db = SessionLocal()
        return db

    @app.teardown_appcontext
    def remove_session(exception=None):
        SessionLocal.remove()

    # --------- Helpers d'authentification ----------

    def current_user():
        if "user_id" not in session:
            return None
        return {
            "id": session.get("user_id"),
            "role": session.get("role"),
            "type": session.get("type"),  # utilisateur ou etudiant
            "nom": session.get("nom"),
        }

    def login_required(role=None):
        def decorator(fn):
            def wrapper(*args, **kwargs):
                user = current_user()
                if not user:
                    flash("Veuillez vous connecter pour accéder à cette page.", "warning")
                    return redirect(url_for("login"))
                if role and user["role"] not in role:
                    flash("Accès non autorisé.", "danger")
                    return redirect(url_for("login"))
                return fn(*args, **kwargs)

            wrapper.__name__ = fn.__name__
            return wrapper

        return decorator

    # --------- Authentification ----------

    @app.route("/", methods=["GET"])
    def index():
        user = current_user()
        if user:
            if user["role"] in ("admin", "enseignant"):
                return redirect(url_for("admin_dashboard"))
            if user["role"] == "etudiant":
                return redirect(url_for("student_dashboard"))
        return redirect(url_for("login"))

    @app.route("/login", methods=["GET", "POST"])
    def login():
        db = get_db()
        if request.method == "POST":
            email = request.form.get("email", "").strip().lower()
            mot_de_passe = request.form.get("mot_de_passe", "").strip()

            # Chercher d'abord dans les utilisateurs (admin/enseignant)
            user = (
                db.query(Utilisateur)
                .filter(Utilisateur.email == email, Utilisateur.mot_de_passe == mot_de_passe)
                .first()
            )
            if user:
                session["user_id"] = user.id
                session["role"] = user.role
                session["type"] = "utilisateur"
                session["nom"] = user.nom
                flash("Connexion réussie.", "success")
                if user.role in ("admin", "enseignant"):
                    return redirect(url_for("admin_dashboard"))

            # Sinon, chercher dans les étudiants
            etu = (
                db.query(Etudiant)
                .filter(Etudiant.email == email, Etudiant.mot_de_passe == mot_de_passe)
                .first()
            )
            if etu:
                session["user_id"] = etu.id
                session["role"] = "etudiant"
                session["type"] = "etudiant"
                session["nom"] = etu.nom
                flash("Connexion réussie.", "success")
                return redirect(url_for("student_dashboard"))

            flash("Identifiants invalides.", "danger")

        return render_template("login.html")

    @app.route("/logout")
    def logout():
        session.clear()
        flash("Vous avez été déconnecté.", "info")
        return redirect(url_for("login"))

    # --------- Tableau de bord Admin ----------

    @app.route("/admin/dashboard")
    @login_required(role=("admin", "enseignant"))
    def admin_dashboard():
        db = get_db()
        total_matieres = db.query(Matiere).count()
        total_devoirs = db.query(Devoir).count()
        total_etudiants = db.query(Etudiant).count()

        moyenne_generale = db.query(func.avg(Note.valeur)).scalar()
        moyenne_generale = round(moyenne_generale, 2) if moyenne_generale is not None else None

        # Statistiques par filière
        etudiants_par_filiere = (
            db.query(Filiere.nom, func.count(Etudiant.id).label("count"))
            .join(Etudiant, Etudiant.filiere_id == Filiere.id, isouter=True)
            .group_by(Filiere.id)
            .all()
        )

        # Statistiques par spécialité
        etudiants_par_specialite = (
            db.query(Specialite.code, func.count(Etudiant.id).label("count"))
            .join(Etudiant, Etudiant.specialite_id == Specialite.id, isouter=True)
            .group_by(Specialite.id)
            .all()
        )

        # Répartition des notes
        all_notes = db.query(Note.valeur).all()
        ranges = {"0-5": 0, "5-10": 0, "10-15": 0, "15-20": 0}
        for (v,) in all_notes:
            if v < 5:
                ranges["0-5"] += 1
            elif v < 10:
                ranges["5-10"] += 1
            elif v < 15:
                ranges["10-15"] += 1
            else:
                ranges["15-20"] += 1
        repartition_notes = [type("Row", (), {"range": k, "count": v})() for k, v in ranges.items() if v > 0]

        # Moyenne par matière
        matieres_difficiles = (
            db.query(Matiere.nom, func.avg(Note.valeur).label("moyenne"))
            .join(Devoir, Devoir.matiere_id == Matiere.id)
            .join(Note, Note.devoir_id == Devoir.id)
            .group_by(Matiere.id)
            .order_by(func.avg(Note.valeur))
            .limit(5)
            .all()
        )

        filieres = db.query(Filiere).order_by(Filiere.nom).all()

        return render_template(
            "admin/dashboard.html",
            total_matieres=total_matieres,
            total_devoirs=total_devoirs,
            total_etudiants=total_etudiants,
            moyenne_generale=moyenne_generale,
            repartition_notes=repartition_notes,
            matieres_difficiles=matieres_difficiles,
            etudiants_par_filiere=etudiants_par_filiere,
            etudiants_par_specialite=etudiants_par_specialite,
            filieres=filieres,
            user=current_user(),
        )

    # --------- CRUD Matières ----------

    @app.route("/admin/matieres")
    @login_required(role=("admin", "enseignant"))
    def admin_matieres():
        db = get_db()
        matieres = db.query(Matiere).order_by(Matiere.nom).all()
        return render_template("admin/matieres.html", matieres=matieres, user=current_user())

    @app.route("/admin/matieres/nouveau", methods=["GET", "POST"])
    @login_required(role=("admin", "enseignant"))
    def admin_matiere_new():
        db = get_db()
        if request.method == "POST":
            code = request.form.get("code")
            nom = request.form.get("nom")
            semestre_id = request.form.get("semestre_id") or None
            filiere_id = request.form.get("filiere_id") or None
            specialite_id = request.form.get("specialite_id") or None
            seuil = float(request.form.get("seuil_validation") or 10)
            rattrapable = 1 if request.form.get("rattrapable") == "on" else 0
            if not code or not nom:
                flash("Le code et le nom de la matière sont obligatoires.", "danger")
            else:
                matiere = Matiere(
                    code=code,
                    nom=nom,
                    semestre_id=int(semestre_id) if semestre_id else None,
                    filiere_id=int(filiere_id) if filiere_id else None,
                    specialite_id=int(specialite_id) if specialite_id else None,
                    seuil_validation=seuil,
                    rattrapable=rattrapable
                )
                db.add(matiere)
                db.commit()
                flash("Matière ajoutée.", "success")
                return redirect(url_for("admin_matieres"))
        semestres = db.query(Semestre).order_by(Semestre.numero).all()
        filieres = db.query(Filiere).order_by(Filiere.nom).all()
        specialites = db.query(Specialite).order_by(Specialite.nom).all()
        return render_template(
            "admin/matiere_form.html",
            matiere=None,
            semestres=semestres,
            filieres=filieres,
            specialites=specialites,
            user=current_user()
        )

    @app.route("/admin/matieres/<int:matiere_id>/edit", methods=["GET", "POST"])
    @login_required(role=("admin", "enseignant"))
    def admin_matiere_edit(matiere_id):
        db = get_db()
        matiere = db.get(Matiere, matiere_id)
        if not matiere:
            flash("Matière introuvable.", "danger")
            return redirect(url_for("admin_matieres"))
        if request.method == "POST":
            matiere.code = request.form.get("code")
            matiere.nom = request.form.get("nom")
            semestre_id = request.form.get("semestre_id") or None
            filiere_id = request.form.get("filiere_id") or None
            specialite_id = request.form.get("specialite_id") or None
            matiere.semestre_id = int(semestre_id) if semestre_id else None
            matiere.filiere_id = int(filiere_id) if filiere_id else None
            matiere.specialite_id = int(specialite_id) if specialite_id else None
            matiere.seuil_validation = float(request.form.get("seuil_validation") or 10)
            matiere.rattrapable = 1 if request.form.get("rattrapable") == "on" else 0
            db.commit()
            flash("Matière mise à jour.", "success")
            return redirect(url_for("admin_matieres"))
        semestres = db.query(Semestre).order_by(Semestre.numero).all()
        filieres = db.query(Filiere).order_by(Filiere.nom).all()
        specialites = db.query(Specialite).order_by(Specialite.nom).all()
        return render_template(
            "admin/matiere_form.html",
            matiere=matiere,
            semestres=semestres,
            filieres=filieres,
            specialites=specialites,
            user=current_user()
        )

    @app.route("/admin/matieres/<int:matiere_id>/delete", methods=["POST"])
    @login_required(role=("admin", "enseignant"))
    def admin_matiere_delete(matiere_id):
        db = get_db()
        matiere = db.get(Matiere, matiere_id)
        if matiere:
            db.delete(matiere)
            db.commit()
            flash("Matière supprimée.", "info")
        return redirect(url_for("admin_matieres"))

    # --------- CRUD Devoirs ----------

    @app.route("/admin/devoirs")
    @login_required(role=("admin", "enseignant"))
    def admin_devoirs():
        db = get_db()
        devoirs = (
            db.query(Devoir)
            .join(Matiere, Devoir.matiere_id == Matiere.id)
            .order_by(Devoir.date.desc())
            .all()
        )
        matieres = db.query(Matiere).order_by(Matiere.nom).all()
        return render_template(
            "admin/devoirs.html",
            devoirs=devoirs,
            matieres=matieres,
            user=current_user(),
        )

    @app.route("/admin/devoirs/nouveau", methods=["GET", "POST"])
    @login_required(role=("admin", "enseignant"))
    def admin_devoir_new():
        db = get_db()
        matieres = db.query(Matiere).order_by(Matiere.nom).all()
        if request.method == "POST":
            nom = request.form.get("nom")
            date_str = request.form.get("date")
            type_devoir = request.form.get("type", "Écrit")
            session_devoir = request.form.get("session", "Normale")
            matiere_id = int(request.form.get("matiere_id"))
            if not nom or not date_str:
                flash("Nom et date du devoir obligatoires.", "danger")
            else:
                dv = Devoir(
                    nom=nom,
                    date=date.fromisoformat(date_str),
                    type=type_devoir,
                    session=session_devoir,
                    matiere_id=matiere_id
                )
                db.add(dv)
                db.commit()
                flash("Devoir ajouté.", "success")
                return redirect(url_for("admin_devoirs"))
        return render_template(
            "admin/devoir_form.html",
            devoir=None,
            matieres=matieres,
            user=current_user(),
        )

    @app.route("/admin/devoirs/<int:devoir_id>/edit", methods=["GET", "POST"])
    @login_required(role=("admin", "enseignant"))
    def admin_devoir_edit(devoir_id):
        db = get_db()
        devoir = db.get(Devoir, devoir_id)
        if not devoir:
            flash("Devoir introuvable.", "danger")
            return redirect(url_for("admin_devoirs"))
        matieres = db.query(Matiere).order_by(Matiere.nom).all()
        if request.method == "POST":
            devoir.nom = request.form.get("nom")
            date_str = request.form.get("date")
            devoir.date = date.fromisoformat(date_str)
            devoir.type = request.form.get("type", "Écrit")
            devoir.session = request.form.get("session", "Normale")
            devoir.matiere_id = int(request.form.get("matiere_id"))
            db.commit()
            flash("Devoir mis à jour.", "success")
            return redirect(url_for("admin_devoirs"))
        return render_template(
            "admin/devoir_form.html",
            devoir=devoir,
            matieres=matieres,
            user=current_user(),
        )

    @app.route("/admin/devoirs/<int:devoir_id>/delete", methods=["POST"])
    @login_required(role=("admin", "enseignant"))
    def admin_devoir_delete(devoir_id):
        db = get_db()
        devoir = db.get(Devoir, devoir_id)
        if devoir:
            db.delete(devoir)
            db.commit()
            flash("Devoir supprimé.", "info")
        return redirect(url_for("admin_devoirs"))

    # --------- CRUD Notes ----------

    @app.route("/admin/notes")
    @login_required(role=("admin", "enseignant"))
    def admin_notes():
        db = get_db()
        notes = (
            db.query(Note)
            .join(Etudiant, Note.etudiant_id == Etudiant.id)
            .join(Devoir, Note.devoir_id == Devoir.id)
            .join(Matiere, Devoir.matiere_id == Matiere.id)
            .order_by(Etudiant.nom)
            .all()
        )
        etudiants = db.query(Etudiant).order_by(Etudiant.nom).all()
        devoirs = (
            db.query(Devoir)
            .join(Matiere, Devoir.matiere_id == Matiere.id)
            .order_by(Matiere.nom, Devoir.date)
            .all()
        )
        return render_template(
            "admin/notes.html",
            notes=notes,
            etudiants=etudiants,
            devoirs=devoirs,
            user=current_user(),
        )

    @app.route("/admin/notes/nouveau", methods=["POST"])
    @login_required(role=("admin", "enseignant"))
    def admin_note_new():
        db = get_db()
        etudiant_id = int(request.form.get("etudiant_id"))
        devoir_id = int(request.form.get("devoir_id"))
        valeur = float(request.form.get("valeur"))

        note = (
            db.query(Note)
            .filter(Note.etudiant_id == etudiant_id, Note.devoir_id == devoir_id)
            .first()
        )
        if note:
            note.valeur = valeur
        else:
            note = Note(etudiant_id=etudiant_id, devoir_id=devoir_id, valeur=valeur)
            db.add(note)
        db.commit()
        flash("Note enregistrée.", "success")
        return redirect(url_for("admin_notes"))

    @app.route("/admin/notes/<int:note_id>/delete", methods=["POST"])
    @login_required(role=("admin", "enseignant"))
    def admin_note_delete(note_id):
        db = get_db()
        note = db.get(Note, note_id)
        if note:
            db.delete(note)
            db.commit()
            flash("Note supprimée.", "info")
        return redirect(url_for("admin_notes"))

    # --------- Gestion des filières ----------

    @app.route("/admin/filieres")
    @login_required(role=("admin",))
    def admin_filieres():
        db = get_db()
        filieres = db.query(Filiere).order_by(Filiere.nom).all()
        return render_template("admin/filieres.html", filieres=filieres, user=current_user())

    @app.route("/admin/filieres/nouveau", methods=["GET", "POST"])
    @login_required(role=("admin",))
    def admin_filiere_new():
        db = get_db()
        if request.method == "POST":
            nom = request.form.get("nom")
            description = request.form.get("description")
            if not nom:
                flash("Le nom de la filière est obligatoire.", "danger")
            else:
                filiere = Filiere(nom=nom, description=description)
                db.add(filiere)
                db.commit()
                flash("Filière ajoutée.", "success")
                return redirect(url_for("admin_filieres"))
        return render_template("admin/filiere_form.html", filiere=None, user=current_user())

    @app.route("/admin/filieres/<int:filiere_id>/delete", methods=["POST"])
    @login_required(role=("admin",))
    def admin_filiere_delete(filiere_id):
        db = get_db()
        filiere = db.get(Filiere, filiere_id)
        if filiere:
            db.delete(filiere)
            db.commit()
            flash("Filière supprimée.", "info")
        return redirect(url_for("admin_filieres"))

    # --------- Gestion des spécialités ----------

    @app.route("/admin/specialites")
    @login_required(role=("admin",))
    def admin_specialites():
        db = get_db()
        specialites = db.query(Specialite).join(Filiere).order_by(Specialite.nom).all()
        return render_template("admin/specialites.html", specialites=specialites, user=current_user())

    @app.route("/admin/specialites/nouveau", methods=["GET", "POST"])
    @login_required(role=("admin",))
    def admin_specialite_new():
        db = get_db()
        if request.method == "POST":
            nom = request.form.get("nom")
            code = request.form.get("code")
            description = request.form.get("description")
            filiere_id = int(request.form.get("filiere_id"))
            if not nom or not code:
                flash("Le nom et le code de la spécialité sont obligatoires.", "danger")
            else:
                specialite = Specialite(nom=nom, code=code, description=description, filiere_id=filiere_id)
                db.add(specialite)
                db.commit()
                flash("Spécialité ajoutée.", "success")
                return redirect(url_for("admin_specialites"))
        filieres = db.query(Filiere).order_by(Filiere.nom).all()
        return render_template("admin/specialite_form.html", specialite=None, filieres=filieres, user=current_user())

    @app.route("/admin/specialites/<int:specialite_id>/delete", methods=["POST"])
    @login_required(role=("admin",))
    def admin_specialite_delete(specialite_id):
        db = get_db()
        specialite = db.get(Specialite, specialite_id)
        if specialite:
            db.delete(specialite)
            db.commit()
            flash("Spécialité supprimée.", "info")
        return redirect(url_for("admin_specialites"))

    # --------- Gestion des semestres ----------

    @app.route("/admin/semestres")
    @login_required(role=("admin",))
    def admin_semestres():
        db = get_db()
        semestres = db.query(Semestre).order_by(Semestre.numero).all()
        return render_template("admin/semestres.html", semestres=semestres, user=current_user())

    @app.route("/admin/semestres/nouveau", methods=["GET", "POST"])
    @login_required(role=("admin",))
    def admin_semestre_new():
        db = get_db()
        if request.method == "POST":
            nom = request.form.get("nom")
            numero = int(request.form.get("numero"))
            if not nom:
                flash("Le nom du semestre est obligatoire.", "danger")
            else:
                semestre = Semestre(nom=nom, numero=numero)
                db.add(semestre)
                db.commit()
                flash("Semestre ajouté.", "success")
                return redirect(url_for("admin_semestres"))
        return render_template("admin/semestre_form.html", semestre=None, user=current_user())

    @app.route("/admin/semestres/<int:semestre_id>/delete", methods=["POST"])
    @login_required(role=("admin",))
    def admin_semestre_delete(semestre_id):
        db = get_db()
        semestre = db.get(Semestre, semestre_id)
        if semestre:
            db.delete(semestre)
            db.commit()
            flash("Semestre supprimé.", "info")
        return redirect(url_for("admin_semestres"))

    # --------- Gestion des utilisateurs ----------

    @app.route("/admin/utilisateurs")
    @login_required(role=("admin",))
    def admin_utilisateurs():
        db = get_db()
        utilisateurs = db.query(Utilisateur).order_by(Utilisateur.nom).all()
        etudiants = db.query(Etudiant).order_by(Etudiant.nom).all()
        return render_template(
            "admin/utilisateurs.html",
            utilisateurs=utilisateurs,
            etudiants=etudiants,
            user=current_user(),
        )

    @app.route("/admin/utilisateurs/nouveau", methods=["GET", "POST"])
    @login_required(role=("admin",))
    def admin_utilisateur_new():
        db = get_db()
        if request.method == "POST":
            nom = request.form.get("nom")
            email = request.form.get("email")
            mot_de_passe = request.form.get("mot_de_passe")
            role = request.form.get("role")

            if role == "etudiant":
                filiere_id = request.form.get("filiere_id") or None
                specialite_id = request.form.get("specialite_id") or None
                etu = Etudiant(
                    nom=nom,
                    email=email,
                    mot_de_passe=mot_de_passe,
                    filiere_id=int(filiere_id) if filiere_id else None,
                    specialite_id=int(specialite_id) if specialite_id else None
                )
                db.add(etu)
            else:
                user = Utilisateur(
                    nom=nom,
                    email=email,
                    mot_de_passe=mot_de_passe,
                    role=role,
                )
                db.add(user)
            db.commit()
            flash("Utilisateur créé.", "success")
            return redirect(url_for("admin_utilisateurs"))
        filieres = db.query(Filiere).order_by(Filiere.nom).all()
        specialites = db.query(Specialite).order_by(Specialite.nom).all()
        return render_template(
            "admin/utilisateur_form.html",
            utilisateur=None,
            filieres=filieres,
            specialites=specialites,
            user=current_user()
        )

    @app.route("/admin/utilisateurs/<string:type>/<int:user_id>/delete", methods=["POST"])
    @login_required(role=("admin",))
    def admin_utilisateur_delete(type, user_id):
        db = get_db()
        if type == "etudiant":
            obj = db.get(Etudiant, user_id)
        else:
            obj = db.get(Utilisateur, user_id)
        if obj:
            db.delete(obj)
            db.commit()
            flash("Utilisateur supprimé.", "info")
        return redirect(url_for("admin_utilisateurs"))

    # --------- Statistiques / Export ----------

    @app.route("/admin/statistiques")
    @login_required(role=("admin", "enseignant"))
    def admin_statistiques():
        db = get_db()

        moyenne_par_etudiant = (
            db.query(Etudiant.nom, func.avg(Note.valeur).label("moyenne"))
            .join(Note, Note.etudiant_id == Etudiant.id)
            .group_by(Etudiant.id)
            .all()
        )
        moyenne_generale = db.query(func.avg(Note.valeur)).scalar()
        moyenne_generale = round(moyenne_generale, 2) if moyenne_generale else None

        # Statut des matières par nombre
        stats_matieres = []
        matieres = db.query(Matiere).all()
        for m in matieres:
            moyenne_matiere = (
                db.query(func.avg(Note.valeur))
                .join(Devoir, Note.devoir_id == Devoir.id)
                .filter(Devoir.matiere_id == m.id)
                .scalar()
            )
            if moyenne_matiere is None:
                statut = "Non notée"
            elif moyenne_matiere >= m.seuil_validation:
                statut = "Validée"
            elif m.rattrapable:
                statut = "À rattraper"
            else:
                statut = "Non validée"
            stats_matieres.append(
                {
                    "nom": m.nom,
                    "moyenne": round(moyenne_matiere, 2) if moyenne_matiere is not None else None,
                    "statut": statut,
                }
            )

        return render_template(
            "admin/statistiques.html",
            moyenne_par_etudiant=moyenne_par_etudiant,
            moyenne_generale=moyenne_generale,
            stats_matieres=stats_matieres,
            user=current_user(),
        )

    # Export CSV/Excel
    @app.route("/admin/export/csv")
    @login_required(role=("admin", "enseignant"))
    def admin_export_csv():
        import csv
        from io import StringIO
        from flask import Response

        db = get_db()
        output = StringIO()
        writer = csv.writer(output, delimiter=";")
        writer.writerow(["Etudiant", "Matière", "Devoir", "Note"])
        rows = (
            db.query(Etudiant.nom, Matiere.nom, Devoir.nom, Note.valeur)
            .join(Note, Note.etudiant_id == Etudiant.id)
            .join(Devoir, Note.devoir_id == Devoir.id)
            .join(Matiere, Devoir.matiere_id == Matiere.id)
            .all()
        )
        for r in rows:
            writer.writerow(r)
        output.seek(0)
        return Response(
            output.read(),
            mimetype="text/csv",
            headers={"Content-Disposition": "attachment;filename=notes.csv"},
        )

    @app.route("/admin/export/excel")
    @login_required(role=("admin", "enseignant"))
    def admin_export_excel():
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from flask import Response
        from io import BytesIO

        db = get_db()
        wb = Workbook()
        ws = wb.active
        ws.title = "Notes"

        # En-têtes
        headers = ["Étudiant", "Filière", "Spécialité", "Matière", "Semestre", "Devoir", "Type", "Date", "Note"]
        ws.append(headers)
        
        # Style des en-têtes
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Données
        rows = (
            db.query(
                Etudiant.nom,
                Filiere.nom,
                Specialite.code,
                Matiere.nom,
                Semestre.nom,
                Devoir.nom,
                Devoir.type,
                Devoir.date,
                Note.valeur
            )
            .join(Note, Note.etudiant_id == Etudiant.id)
            .join(Devoir, Note.devoir_id == Devoir.id)
            .join(Matiere, Devoir.matiere_id == Matiere.id)
            .outerjoin(Filiere, Etudiant.filiere_id == Filiere.id)
            .outerjoin(Specialite, Etudiant.specialite_id == Specialite.id)
            .outerjoin(Semestre, Matiere.semestre_id == Semestre.id)
            .order_by(Etudiant.nom, Matiere.nom)
            .all()
        )
        for r in rows:
            ws.append([
                r[0],  # Étudiant
                r[1] or "-",  # Filière
                r[2] or "-",  # Spécialité
                r[3],  # Matière
                r[4] or "-",  # Semestre
                r[5],  # Devoir
                r[6],  # Type
                r[7].strftime("%d/%m/%Y") if r[7] else "-",  # Date
                r[8]  # Note
            ])

        # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return Response(
            output.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment;filename=notes_completes.xlsx"},
        )

    @app.route("/admin/export/excel/filiere/<int:filiere_id>")
    @login_required(role=("admin", "enseignant"))
    def admin_export_excel_filiere(filiere_id):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from flask import Response
        from io import BytesIO

        db = get_db()
        filiere = db.get(Filiere, filiere_id)
        if not filiere:
            flash("Filière introuvable.", "danger")
            return redirect(url_for("admin_statistiques"))

        wb = Workbook()
        ws = wb.active
        ws.title = f"Notes {filiere.nom}"

        # En-têtes
        headers = ["Étudiant", "Spécialité", "Matière", "Semestre", "Devoir", "Type", "Date", "Note"]
        ws.append(headers)
        
        # Style
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Données filtrées par filière
        rows = (
            db.query(
                Etudiant.nom,
                Specialite.code,
                Matiere.nom,
                Semestre.nom,
                Devoir.nom,
                Devoir.type,
                Devoir.date,
                Note.valeur
            )
            .join(Note, Note.etudiant_id == Etudiant.id)
            .join(Devoir, Note.devoir_id == Devoir.id)
            .join(Matiere, Devoir.matiere_id == Matiere.id)
            .filter(Etudiant.filiere_id == filiere_id)
            .outerjoin(Specialite, Etudiant.specialite_id == Specialite.id)
            .outerjoin(Semestre, Matiere.semestre_id == Semestre.id)
            .order_by(Etudiant.nom, Matiere.nom)
            .all()
        )
        for r in rows:
            ws.append([
                r[0], r[1] or "-", r[2], r[3] or "-", r[4], r[5],
                r[6].strftime("%d/%m/%Y") if r[6] else "-", r[7]
            ])

        # Ajuster largeur
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return Response(
            output.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment;filename=notes_{filiere.nom}.xlsx"},
        )

    @app.route("/admin/export/excel/etudiant/<int:etudiant_id>")
    @login_required(role=("admin", "enseignant"))
    def admin_export_excel_etudiant(etudiant_id):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from flask import Response
        from io import BytesIO

        db = get_db()
        etudiant = db.get(Etudiant, etudiant_id)
        if not etudiant:
            flash("Étudiant introuvable.", "danger")
            return redirect(url_for("admin_utilisateurs"))

        wb = Workbook()
        ws = wb.active
        ws.title = "Relevé de notes"

        # En-têtes
        ws.append([f"Relevé de notes - {etudiant.nom}"])
        ws.append([f"Filière: {etudiant.filiere.nom if etudiant.filiere else '-'}"])
        ws.append([f"Spécialité: {etudiant.specialite.nom if etudiant.specialite else '-'}"])
        ws.append([])
        
        headers = ["Matière", "Semestre", "Devoir", "Type", "Date", "Note"]
        ws.append(headers)
        
        # Style
        for cell in ws[5]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Données
        rows = (
            db.query(
                Matiere.nom,
                Semestre.nom,
                Devoir.nom,
                Devoir.type,
                Devoir.date,
                Note.valeur
            )
            .join(Note, Note.devoir_id == Devoir.id)
            .join(Devoir, Note.devoir_id == Devoir.id)
            .join(Matiere, Devoir.matiere_id == Matiere.id)
            .filter(Note.etudiant_id == etudiant_id)
            .outerjoin(Semestre, Matiere.semestre_id == Semestre.id)
            .order_by(Matiere.nom, Devoir.date)
            .all()
        )
        for r in rows:
            ws.append([
                r[0], r[1] or "-", r[2], r[3],
                r[4].strftime("%d/%m/%Y") if r[4] else "-", r[5]
            ])

        # Ajuster largeur
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return Response(
            output.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment;filename=releve_{etudiant.nom.replace(' ', '_')}.xlsx"},
        )

    # --------- Interface Étudiant ----------

    def calcul_moyennes_etudiant(db, etudiant_id):
        matieres = db.query(Matiere).all()
        resultats = []
        for m in matieres:
            moyenne = (
                db.query(func.avg(Note.valeur))
                .join(Devoir, Note.devoir_id == Devoir.id)
                .filter(Devoir.matiere_id == m.id, Note.etudiant_id == etudiant_id)
                .scalar()
            )
            if moyenne is None:
                statut = "Non notée"
            elif moyenne >= m.seuil_validation:
                statut = "Validée"
            elif m.rattrapable:
                statut = "À rattraper"
            else:
                statut = "Non validée"
            resultats.append(
                {
                    "matiere": m,
                    "moyenne": round(moyenne, 2) if moyenne is not None else None,
                    "statut": statut,
                }
            )
        # Moyenne générale
        moyenne_generale = (
            db.query(func.avg(Note.valeur)).filter(Note.etudiant_id == etudiant_id).scalar()
        )
        moyenne_generale = round(moyenne_generale, 2) if moyenne_generale else None
        return resultats, moyenne_generale

    @app.route("/etudiant/dashboard")
    @login_required(role=("etudiant",))
    def student_dashboard():
        db = get_db()
        user = current_user()
        etudiant = db.get(Etudiant, user["id"])
        resultats, moyenne_generale = calcul_moyennes_etudiant(db, user["id"])

        # Devoirs manquants: devoirs sans note pour cet étudiant
        devoirs_tous = db.query(Devoir).all()
        devoirs_notes_ids = {
            n.devoir_id for n in db.query(Note).filter(Note.etudiant_id == user["id"]).all()
        }
        devoirs_manquants = [d for d in devoirs_tous if d.id not in devoirs_notes_ids]

        return render_template(
            "etudiant/dashboard.html",
            etudiant=etudiant,
            resultats=resultats,
            moyenne_generale=moyenne_generale,
            devoirs_manquants=devoirs_manquants,
            user=user,
        )

    @app.route("/etudiant/matieres")
    @login_required(role=("etudiant",))
    def student_matieres():
        db = get_db()
        user = current_user()
        resultats, moyenne_generale = calcul_moyennes_etudiant(db, user["id"])
        return render_template(
            "etudiant/matieres.html",
            resultats=resultats,
            moyenne_generale=moyenne_generale,
            user=user,
        )

    @app.route("/etudiant/devoirs")
    @login_required(role=("etudiant",))
    def student_devoirs():
        db = get_db()
        user = current_user()
        matiere_id = request.args.get("matiere_id", type=int)

        query = (
            db.query(Devoir, Note.valeur)
            .join(Matiere, Devoir.matiere_id == Matiere.id)
            .outerjoin(
                Note,
                (Note.devoir_id == Devoir.id) & (Note.etudiant_id == user["id"]),
            )
        )
        if matiere_id:
            query = query.filter(Devoir.matiere_id == matiere_id)
        devoirs = query.order_by(Devoir.date).all()
        matieres = db.query(Matiere).order_by(Matiere.nom).all()
        return render_template(
            "etudiant/devoirs.html",
            devoirs=devoirs,
            matieres=matieres,
            matiere_id=matiere_id,
            user=user,
        )

    @app.route("/etudiant/resultats")
    @login_required(role=("etudiant",))
    def student_resultats():
        db = get_db()
        user = current_user()
        resultats, moyenne_generale = calcul_moyennes_etudiant(db, user["id"])
        
        # Récupérer tous les semestres
        semestres = db.scalars(select(Semestre).order_by(Semestre.numero)).all()
        
        # Grouper les résultats par semestre
        resultats_par_semestre = {}
        for r in resultats:
            if r["matiere"].semestre:
                sem_id = r["matiere"].semestre.id
                if sem_id not in resultats_par_semestre:
                    resultats_par_semestre[sem_id] = []
                resultats_par_semestre[sem_id].append(r)
        
        return render_template(
            "etudiant/resultats.html",
            resultats=resultats,
            moyenne_generale=moyenne_generale,
            semestres=semestres,
            resultats_par_semestre=resultats_par_semestre,
            user=user,
        )

    @app.route("/etudiant/export/excel")
    @login_required(role=("etudiant",))
    def student_export_excel():
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from flask import Response
        from io import BytesIO

        db = get_db()
        user = current_user()
        etudiant = db.get(Etudiant, user["id"])
        
        # Récupérer le semestre filtré
        semestre_id = request.args.get("semestre_id", type=int)
        
        wb = Workbook()
        ws = wb.active
        
        if semestre_id:
            semestre = db.get(Semestre, semestre_id)
            ws.title = f"Notes {semestre.nom if semestre else 'Semestre'}"
            titre_semestre = f" - {semestre.nom}" if semestre else ""
        else:
            ws.title = "Mes notes"
            titre_semestre = ""

        # Informations étudiant
        ws.append([f"Relevé de notes{titre_semestre} - {etudiant.nom}"])
        ws.append([f"Filière: {etudiant.filiere.nom if etudiant.filiere else '-'}"])
        ws.append([f"Spécialité: {etudiant.specialite.nom if etudiant.specialite else '-'}"])
        ws.append([])
        
        headers = ["Matière", "Code", "Semestre", "Devoir", "Type", "Date", "Note"]
        ws.append(headers)
        
        # Style
        for cell in ws[5]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Données filtrées
        query = (
            db.query(
                Matiere.nom,
                Matiere.code,
                Semestre.nom,
                Devoir.nom,
                Devoir.type,
                Devoir.date,
                Note.valeur
            )
            .select_from(Note)
            .join(Devoir, Note.devoir_id == Devoir.id)
            .join(Matiere, Devoir.matiere_id == Matiere.id)
            .outerjoin(Semestre, Matiere.semestre_id == Semestre.id)
            .filter(Note.etudiant_id == user["id"])
        )
        
        if semestre_id:
            query = query.filter(Matiere.semestre_id == semestre_id)
            
        rows = query.order_by(Matiere.nom, Devoir.date).all()
        
        for r in rows:
            ws.append([
                r[0], r[1], r[2] or "-", r[3], r[4],
                r[5].strftime("%d/%m/%Y") if r[5] else "-", r[6]
            ])

        # Ajuster largeur
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"notes_{semestre.nom if semestre_id and semestre else 'completes'}_{etudiant.nom.replace(' ', '_')}.xlsx"
        
        return Response(
            output.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment;filename={filename}"},
        )

    return app


app = create_app()

if __name__ == "__main__":
    app.run(host='0.0.0.0', port=5000, debug=False)

