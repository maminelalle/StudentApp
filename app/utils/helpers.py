"""
Utilitaires pour l'application SupNum Share
Inclut les fonctions d'export, calcul de moyennes, etc.
"""

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
import os


class MoyenneCalculator:
    """Calcule les moyennes et les statuts"""
    
    @staticmethod
    def calculate_matiere_moyenne(notes):
        """Calcule la moyenne pondérée d'une matière"""
        if not notes:
            return None
        
        total_coeff = sum(n.devoir.coefficient for n in notes if n.devoir)
        if total_coeff == 0:
            return None
        
        total = sum(n.valeur * n.devoir.coefficient for n in notes if n.devoir)
        return round(total / total_coeff, 2)
    
    @staticmethod
    def calculate_semestre_moyenne(notes_semestre, matieres_semestre):
        """Calcule la moyenne du semestre"""
        if not notes_semestre or not matieres_semestre:
            return None
        
        moyennes_matieres = {}
        for matiere in matieres_semestre:
            notes = [n for n in notes_semestre if n.matiere_id == matiere.id]
            if notes:
                moyennes_matieres[matiere.id] = {
                    'moyenne': MoyenneCalculator.calculate_matiere_moyenne(notes),
                    'coefficient': matiere.coefficient
                }
        
        if not moyennes_matieres:
            return None
        
        total_coeff = sum(m['coefficient'] for m in moyennes_matieres.values() if m['moyenne'] is not None)
        if total_coeff == 0:
            return None
        
        total = sum(m['moyenne'] * m['coefficient'] for m in moyennes_matieres.values() if m['moyenne'] is not None)
        return round(total / total_coeff, 2)
    
    @staticmethod
    def get_statut(moyenne):
        """Retourne le statut en fonction de la moyenne"""
        if moyenne is None:
            return "Non évalué"
        elif moyenne >= 10:
            return "Validée"
        elif moyenne >= 8:
            return "À rattrapage"
        else:
            return "Non validée"
    
    @staticmethod
    def get_statut_color(statut):
        """Retourne la couleur associée au statut"""
        colors_map = {
            "Validée": "green",
            "À rattrapage": "orange",
            "Non validée": "red",
            "Non évalué": "gray"
        }
        return colors_map.get(statut, "gray")


class StatisticsCalculator:
    """Calcule les statistiques pour les tableaux de bord"""
    
    @staticmethod
    def calculate_student_stats(etudiant, notes):
        """Calcule les statistiques d'un étudiant"""
        matieres_validees = 0
        matieres_rattrapage = 0
        matieres_echouees = 0
        
        for matiere in etudiant.filiere.matieres:
            notes_matiere = [n for n in notes if n.matiere_id == matiere.id]
            if notes_matiere:
                moyenne = MoyenneCalculator.calculate_matiere_moyenne(notes_matiere)
                statut = MoyenneCalculator.get_statut(moyenne)
                
                if statut == "Validée":
                    matieres_validees += 1
                elif statut == "À rattrapage":
                    matieres_rattrapage += 1
                elif statut == "Non validée":
                    matieres_echouees += 1
        
        return {
            'validees': matieres_validees,
            'rattrapage': matieres_rattrapage,
            'echouees': matieres_echouees,
            'total': matieres_validees + matieres_rattrapage + matieres_echouees
        }


class ExcelExporter:
    """Exporte les données en format Excel"""
    
    @staticmethod
    def export_student_notes(etudiant, notes, filename=None):
        """Exporte les notes d'un étudiant en Excel"""
        if filename is None:
            filename = f'notes_{etudiant.user.username}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Notes"
        
        # En-tête
        ws['A1'] = "RELEVÉ DE NOTES"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:E1')
        
        ws['A2'] = f"Étudiant: {etudiant.user.prenom} {etudiant.user.nom}"
        ws['A3'] = f"Filière: {etudiant.filiere.nom}"
        ws['A4'] = f"Spécialité: {etudiant.specialite.nom}"
        
        # Titres des colonnes
        ws['A6'] = "Matière"
        ws['B6'] = "Code"
        ws['C6'] = "Coefficient"
        ws['D6'] = "Moyenne"
        ws['E6'] = "Statut"
        
        # Style en-têtes
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}6'].fill = header_fill
            ws[f'{col}6'].font = header_font
        
        # Données
        row = 7
        for matiere in etudiant.filiere.matieres:
            notes_matiere = [n for n in notes if n.matiere_id == matiere.id]
            if notes_matiere:
                moyenne = MoyenneCalculator.calculate_matiere_moyenne(notes_matiere)
                statut = MoyenneCalculator.get_statut(moyenne)
                
                ws[f'A{row}'] = matiere.nom
                ws[f'B{row}'] = matiere.code
                ws[f'C{row}'] = matiere.coefficient
                ws[f'D{row}'] = moyenne if moyenne else "N/A"
                ws[f'E{row}'] = statut
                
                row += 1
        
        # Ajuster les largeurs
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        
        return wb, filename


class PDFExporter:
    """Exporte les données en format PDF"""
    
    @staticmethod
    def export_student_notes(etudiant, notes, filename=None):
        """Exporte les notes d'un étudiant en PDF"""
        if filename is None:
            filename = f'notes_{etudiant.user.username}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        
        # Créer le document PDF
        doc = SimpleDocTemplate(
            filename,
            pagesize=A4,
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.75*inch,
            bottomMargin=0.5*inch,
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Titre
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#2563EB'),
            spaceAfter=30,
            alignment=1,  # Center
        )
        
        elements.append(Paragraph("RELEVÉ DE NOTES OFFICIEL", title_style))
        elements.append(Spacer(1, 12))
        
        # Informations étudiant
        info_text = f"""
        <b>Étudiant:</b> {etudiant.user.prenom} {etudiant.user.nom}<br/>
        <b>Filière:</b> {etudiant.filiere.nom}<br/>
        <b>Spécialité:</b> {etudiant.specialite.nom}<br/>
        <b>Date d'édition:</b> {datetime.now().strftime('%d/%m/%Y à %H:%M')}<br/>
        """
        elements.append(Paragraph(info_text, styles['Normal']))
        elements.append(Spacer(1, 12))
        
        # Tableau des notes
        data = [['Matière', 'Code', 'Coefficient', 'Moyenne', 'Statut']]
        
        for matiere in etudiant.filiere.matieres:
            notes_matiere = [n for n in notes if n.matiere_id == matiere.id]
            if notes_matiere:
                moyenne = MoyenneCalculator.calculate_matiere_moyenne(notes_matiere)
                statut = MoyenneCalculator.get_statut(moyenne)
                
                data.append([
                    matiere.nom,
                    matiere.code,
                    str(matiere.coefficient),
                    str(moyenne) if moyenne else "N/A",
                    statut
                ])
        
        table = Table(data, colWidths=[2.5*inch, 1*inch, 1*inch, 1*inch, 1.2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')]),
        ]))
        
        elements.append(table)
        
        # Construire le PDF
        doc.build(elements)
        
        return filename


def allowed_file(filename):
    """Vérifie si le fichier est autorisé"""
    from app import ALLOWED_EXTENSIONS
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def format_file_size(size_bytes):
    """Formate la taille du fichier en format lisible"""
    for unit in ['B', 'KB', 'MB', 'GB']:
        if size_bytes < 1024.0:
            return f"{size_bytes:.2f} {unit}"
        size_bytes /= 1024.0
    return f"{size_bytes:.2f} TB"


__all__ = [
    'MoyenneCalculator',
    'StatisticsCalculator',
    'ExcelExporter',
    'PDFExporter',
    'allowed_file',
    'format_file_size',
]
